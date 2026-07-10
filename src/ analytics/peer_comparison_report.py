import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ---------------------------------------------------
# Project Paths
# ---------------------------------------------------

project_root = Path(__file__).resolve().parents[2]

db_path = project_root / "db" / "nifty100.db"

peer_file = project_root / "data" / "processed" / "peer_groups_cleaned.csv"

output_path = project_root / "output"

output_path.mkdir(exist_ok=True)

excel_file = output_path / "peer_comparison.xlsx"

# ---------------------------------------------------
# Database Connection
# ---------------------------------------------------

conn = sqlite3.connect(db_path)

print("Database Connected")

# ---------------------------------------------------
# Load Financial Ratios
# ---------------------------------------------------

financial = pd.read_sql(
    "SELECT * FROM financial_ratios",
    conn
)

print("Financial Ratios Loaded")

# ---------------------------------------------------
# Load Peer Groups
# ---------------------------------------------------

peer = pd.read_csv(
    peer_file,
    header=None
)

peer.columns = [
    "id",
    "peer_group_name",
    "company_id",
    "is_active"
]

print("Peer Groups Loaded")

# ---------------------------------------------------
# Merge Tables
# ---------------------------------------------------

df = pd.merge(
    financial,
    peer,
    on="company_id",
    how="left"
)

df = df.dropna(subset=["peer_group_name"])

print("Merged Shape :", df.shape)

# ---------------------------------------------------
# Workbook
# ---------------------------------------------------

wb = Workbook()

# Remove default sheet

wb.remove(wb.active)

# ---------------------------------------------------
# Excel Styles
# ---------------------------------------------------

header_fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

center = Alignment(
    horizontal="center"
)

green_fill = PatternFill(
    start_color="92D050",
    end_color="92D050",
    fill_type="solid"
)

yellow_fill = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="F4CCCC",
    end_color="F4CCCC",
    fill_type="solid"
)

gold_fill = PatternFill(
    start_color="FFC000",
    end_color="FFC000",
    fill_type="solid"
)

print("Workbook Created")

# ---------------------------------------------------
# Metrics to Export
# ---------------------------------------------------

metrics = [

    "roe_calculated",
    "roa",
    "net_profit_margin",
    "operating_profit_margin",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow",
    "cfo_quality",
    "capex_intensity",
    "fcf_conversion",
    "revenue_cagr_3yr",
    "profit_cagr_3yr",
    "eps_cagr_3yr",
    "revenue_cagr_5yr",
    "profit_cagr_5yr",
    "eps_cagr_5yr",
    "revenue_cagr_10yr",
    "profit_cagr_10yr",
    "eps_cagr_10yr"

]

peer_groups = sorted(df["peer_group_name"].unique())

print("Peer Groups :", len(peer_groups))

# ---------------------------------------------------
# Create One Sheet Per Peer Group
# ---------------------------------------------------

for group in peer_groups:

    print(f"Processing : {group}")

    sheet = wb.create_sheet(title=group[:31])

    peer_df = df[
        df["peer_group_name"] == group
    ].copy()

    columns = ["company_id"] + metrics

    peer_df = peer_df[columns]

    # ------------------------------------------
    # Header
    # ------------------------------------------

    for col_num, column_name in enumerate(columns, start=1):

        cell = sheet.cell(
            row=1,
            column=col_num
        )

        cell.value = column_name

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = center

    # ------------------------------------------
    # Write Company Data
    # ------------------------------------------

    for row_num, (_, row) in enumerate(peer_df.iterrows(), start=2):

        for col_num, column_name in enumerate(columns, start=1):

            value = row[column_name]

            sheet.cell(
                row=row_num,
                column=col_num
            ).value = value

    print(
        f"Rows Written : {len(peer_df)}"
    )
    # ---------------------------------------------------
    # Apply Formatting
    # ---------------------------------------------------

    for sheet in wb.worksheets:

        max_row = sheet.max_row
        max_col = sheet.max_column

        # ------------------------------------------
        # Auto Width
        # ------------------------------------------

        for column in sheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            sheet.column_dimensions[column_letter].width = max_length + 3

        # ------------------------------------------
        # Colour Formatting
        # ------------------------------------------

        for row in range(2, max_row + 1):

            for col in range(2, max_col + 1):

                cell = sheet.cell(row=row, column=col)

                value = cell.value

                if isinstance(value, (int, float)):

                    if value >= 75:

                        cell.fill = green_fill

                    elif value >= 25:

                        cell.fill = yellow_fill

                    else:

                        cell.fill = red_fill

        # ------------------------------------------
        # Highlight Benchmark Company
        # ------------------------------------------

        for row in range(2, max_row + 1):

            benchmark = sheet.cell(row=row, column=max_col).value

            if benchmark == 1 or benchmark is True:

                for col in range(1, max_col + 1):
                    sheet.cell(row=row, column=col).fill = gold_fill

        # ------------------------------------------
        # Summary Row
        # ------------------------------------------

        summary_row = max_row + 2

        sheet.cell(summary_row, 1).value = "Median"

        sheet.cell(summary_row, 1).font = Font(bold=True)

        for col in range(2, max_col + 1):

            values = []

            for row in range(2, max_row + 1):

                value = sheet.cell(row=row, column=col).value

                if isinstance(value, (int, float)):
                    values.append(value)

            if len(values) > 0:
                sheet.cell(summary_row, col).value = round(

                    pd.Series(values).median(),

                    2

                )

    # ---------------------------------------------------
    # Save Workbook
    # ---------------------------------------------------

    wb.save(excel_file)

    conn.close()

    print("=" * 60)
    print("DAY 20 COMPLETED SUCCESSFULLY")
    print("=" * 60)

    print("Workbook Saved Successfully")
    print(excel_file)

    print("Sheets Created :", len(wb.sheetnames))

    print("Sheet Names:")

    for sheet in wb.sheetnames:
        print("-", sheet)

    print("=" * 60)