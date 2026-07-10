from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

project_root = Path(__file__).resolve().parents[2]

excel = project_root/"output"/"screener_output.xlsx"

wb = load_workbook(excel)

ws = wb["Composite Score"]

green = PatternFill(

    fill_type="solid",

    start_color="90EE90"

)

red = PatternFill(

    fill_type="solid",

    start_color="FFC7CE"

)

score_col = None

for cell in ws[1]:
    if cell.value == "composite_score":
        score_col = cell.column
        break

if score_col:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=score_col)
        if isinstance(cell.value, (int, float)):
            if cell.value >= 70:
                cell.fill = green
            else:
                cell.fill = red

wb.save(excel)

print("Excel Formatting Completed")